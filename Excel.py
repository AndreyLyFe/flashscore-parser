from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A1'] = 'Дата'
ws['B1'] = 'Время'

ws['E1'] = 'Команда 2'
ws['D1'] = 'Команда 1'


ws['H1'] = 'П1, %'
ws['F1'] = 'Был'
ws['G1'] = "Стал"
ws['I1'] = "Шанс, %"
ws['c1'] = 'Лига'
ws['M1'] = 'Ничья, %'
ws['K1'] = 'Был'
ws['L1'] = "Стал"
ws['N1'] = "Шанс, %"

ws['R1'] = 'П2, %'
ws['P1'] = 'Был'
ws['Q1'] = "Стал"
ws['S1'] = "Шанс, %"

ws['U1'] = 'Предположение'
ws['V1'] = 'Тотал'
ws['W1'] = 'Ключ'